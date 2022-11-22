from asyncio.windows_events import NULL
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm
from openpyxl.utils import get_column_letter

workbook = load_workbook(filename='LAB_DYPLOMOWE/bisz2.xlsx')
sheet = workbook.active

for i in tqdm(range(1, sheet.max_row+1)):
    for j in range(1, 5):
        char = get_column_letter(j)
        if sheet[char + str(i)].value == '             ':
            sheet[char + str(i)] = sheet[char + str(i+1)].value
    # for j in range(1, 5):
    #     char = get_column_letter(j)    
    #     if sheet[char + str(i)].value == 'kwak':
    #         if sheet[char + str(i+1)].value == 'kwak':
    #             sheet[char + str(i)] = sheet[char + str(i+2)].value
    #         else:
    #             sheet[char + str(i)] = sheet[char + str(i+1)].value
    #     else: 
    #         pass

    for j in range(1,5):
        char = get_column_letter(j)
        if sheet[char + str(i)].value == 'kwak':
            sheet[char + str(i)] = sheet[char + str(i+1)].value

# print('.' + sheet["C51"].value + '.')

workbook.save(filename='LAB_DYPLOMOWE/bisz3.xlsx')